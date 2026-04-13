# -*- coding: utf-8 -*-
"""Excel 出力。"""

from __future__ import annotations

from numbers import Number
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class ExportError(Exception):
    """Excel 出力エラー。"""


CHART_STYLES = {
    "実績": {"color": "2F75B5", "marker": "circle", "dashed": False},
    "直線延長予測": {"color": "C2410C", "marker": "square", "dashed": True},
    "重み付き回帰予測": {"color": "EA580C", "marker": "diamond", "dashed": True},
    "予測": {"color": "ED7D31", "marker": "square", "dashed": True},
    "外部要因予測": {"color": "70AD47", "marker": "triangle", "dashed": False},
}

GROUP_FILLS = {
    "実績": PatternFill("solid", fgColor="EAF3FF"),
    "直線": PatternFill("solid", fgColor="FFF5EC"),
    "重み": PatternFill("solid", fgColor="FFF1E6"),
    "AI": PatternFill("solid", fgColor="ECF9F0"),
}

HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
TITLE_FONT = Font(size=13, bold=True)
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _prepare_output_path(file_path: str) -> Path:
    path = Path(file_path)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise ExportError(f"出力先フォルダを作成できません: {e}") from e
    return path


def export_dataframe(
    file_path: str,
    df: pd.DataFrame,
    sheet_name: str = "一覧",
    table_name: Optional[str] = None,
    yearly_chart_df: Optional[pd.DataFrame] = None,
    chart_title: str = "年別推移",
    chart_subtitle: str = "",
) -> None:
    path = _prepare_output_path(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet"

    row = 1
    if table_name:
        ws.cell(row=row, column=1, value=table_name)
        ws.cell(row=row, column=1).font = TITLE_FONT
        row += 1

    row = _write_dataframe(ws, df, start_row=row)
    _auto_fit_columns(ws)

    if yearly_chart_df is not None and not yearly_chart_df.empty:
        chart_source = _build_chart_source_from_long(yearly_chart_df)
        if not chart_source.empty:
            ws_src = wb.create_sheet("グラフ元データ")
            _write_dataframe(ws_src, chart_source, start_row=1)
            ws_src.sheet_state = "hidden"

            ws_chart = wb.create_sheet("グラフ")
            ws_chart["A1"] = chart_title
            ws_chart["A1"].font = TITLE_FONT
            if chart_subtitle:
                ws_chart["A2"] = chart_subtitle
            _add_long_chart_pair(
                ws_chart,
                ws_src,
                chart_source,
                start_row=4 if chart_subtitle else 3,
            )

    _save_workbook(wb, path)


def export_forecast_workbook(
    file_path: str,
    comparison_df: pd.DataFrame,
    *,
    meta_lines: list[str] | None = None,
    chart_subtitle: str = "",
) -> None:
    path = _prepare_output_path(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "年次予測"

    row = 1
    ws.cell(row=row, column=1, value="顧客別納入分析システム / 年次予測")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 1

    if chart_subtitle:
        ws.cell(row=row, column=1, value=chart_subtitle)
        row += 1

    for line in meta_lines or []:
        ws.cell(row=row, column=1, value=f"・{line}")
        row += 1

    row += 1
    table_start = row
    table_end = _write_dataframe(ws, comparison_df, start_row=table_start)
    _style_forecast_table(ws, table_start, table_end, len(comparison_df.columns))
    _auto_fit_columns(ws, max_col=len(comparison_df.columns))

    ws_src = wb.create_sheet("予測グラフ元データ")
    _write_forecast_chart_source(ws_src, comparison_df)
    ws_src.sheet_state = "hidden"

    _add_comparison_chart_pair(
        ws,
        ws_src,
        comparison_df,
        start_row=table_end + 3,
    )

    _save_workbook(wb, path)


def _write_dataframe(ws, df: Optional[pd.DataFrame], *, start_row: int) -> int:
    row_cursor = start_row
    if df is None or df.empty:
        ws.cell(row=row_cursor, column=1, value="データがありません。")
        return row_cursor

    columns = [str(col) for col in df.columns]
    for row in dataframe_to_rows(df, index=False, header=True):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if row_cursor == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws.row_dimensions[row_cursor].height = 30
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                header_name = columns[col_idx - 1] if col_idx - 1 < len(columns) else ""
                if (
                    header_name not in ("年", "月")
                    and isinstance(value, Number)
                    and pd.notna(value)
                ):
                    cell.number_format = "#,##0"
        row_cursor += 1
    return row_cursor - 1


def _build_chart_source_from_long(yearly_df: pd.DataFrame) -> pd.DataFrame:
    work = yearly_df.copy()
    if work.empty:
        return pd.DataFrame()
    if "種別" not in work.columns:
        work["種別"] = "実績"
    work["年"] = pd.to_numeric(work["年"], errors="coerce")
    work["納品数"] = pd.to_numeric(work["納品数"], errors="coerce")
    work["金額"] = pd.to_numeric(work["金額"], errors="coerce")
    work = work.dropna(subset=["年"]).copy()
    work["年"] = work["年"].astype(int)

    pivot = work.pivot_table(
        index="年",
        columns="種別",
        values=["納品数", "金額"],
        aggfunc="sum",
    )
    if pivot.empty:
        return pd.DataFrame()
    pivot.columns = [f"{kind}_{metric}" for metric, kind in pivot.columns.to_flat_index()]
    ordered_cols = [
        f"{kind}_{metric}"
        for kind in ("実績", "直線延長予測", "重み付き回帰予測", "外部要因予測", "予測")
        for metric in ("納品数", "金額")
        if f"{kind}_{metric}" in pivot.columns
    ]
    return pivot.reset_index()[["年", *ordered_cols]]


def _configure_line_series(series, *, color: str, marker: str, dashed: bool) -> None:
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = 18000
    if dashed:
        series.graphicalProperties.line.dashStyle = "sysDot"
    series.marker.symbol = marker
    series.marker.size = 6
    series.graphicalProperties.solidFill = color


def _build_line_chart(*, title: str, y_title: str) -> LineChart:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "年（西暦）"
    chart.legend.position = "r"
    chart.height = 10.5
    chart.width = 21.0
    chart.style = 2
    chart.y_axis.number_format = "#,##0"
    chart.x_axis.number_format = "0"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.minorTickMark = "none"
    chart.y_axis.minorTickMark = "none"
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "autoZero"
    chart.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget="inner",
            xMode="edge",
            yMode="edge",
            x=0.16,
            y=0.08,
            w=0.62,
            h=0.62,
        )
    )
    return chart


def _add_long_chart_pair(ws_chart, ws_src, chart_source: pd.DataFrame, *, start_row: int) -> None:
    row_count = len(chart_source.index) + 1
    qty_chart = _build_line_chart(title="納品数", y_title="納品数（年合計）")
    amount_chart = _build_line_chart(title="金額", y_title="金額（円・年合計）")

    categories = Reference(ws_src, min_col=1, min_row=2, max_row=row_count)
    for idx, column in enumerate(chart_source.columns[1:], start=2):
        values = Reference(ws_src, min_col=idx, min_row=1, max_row=row_count)
        series = qty_chart.series if "納品数" in column else amount_chart.series
        target_chart = qty_chart if "納品数" in column else amount_chart
        target_chart.add_data(values, titles_from_data=True)
        target_chart.set_categories(categories)
        style_key = str(column).split("_", 1)[0]
        style = CHART_STYLES.get(style_key, {"color": "64748B", "marker": "circle", "dashed": False})
        _configure_line_series(
            target_chart.series[-1],
            color=style["color"],
            marker=style["marker"],
            dashed=style["dashed"],
        )

    ws_chart.add_chart(qty_chart, f"A{start_row}")
    ws_chart.add_chart(amount_chart, f"A{start_row + 18}")


def _style_forecast_table(ws, start_row: int, end_row: int, end_col: int) -> None:
    for col_idx in range(1, end_col + 1):
        header = str(ws.cell(row=start_row, column=col_idx).value or "")
        fill = None
        if header.startswith("実績(") or header.startswith("実績\n"):
            fill = GROUP_FILLS["実績"]
        elif header.startswith("直線(") or header.startswith("直線延長\n"):
            fill = GROUP_FILLS["直線"]
        elif header.startswith("重み(") or header.startswith("重み付き回帰\n"):
            fill = GROUP_FILLS["重み"]
        elif header.startswith("外部(") or header.startswith("外部要因予測\n"):
            fill = GROUP_FILLS["AI"]

        if fill is None:
            continue
        for row_idx in range(start_row, end_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _add_comparison_chart_pair(
    ws,
    ws_src,
    chart_df: pd.DataFrame,
    *,
    start_row: int,
) -> None:
    if chart_df is None or chart_df.empty:
        return

    row_count = len(chart_df.index)
    qty_title_row = start_row - 1
    qty_chart_row = start_row
    amount_title_row = start_row + 22
    amount_chart_row = start_row + 23

    ws.cell(row=qty_title_row, column=1, value="納品数推移")
    ws.cell(row=qty_title_row, column=1).font = HEADER_FONT
    ws.cell(row=amount_title_row, column=1, value="金額推移")
    ws.cell(row=amount_title_row, column=1).font = HEADER_FONT

    qty_chart = _build_line_chart(title="納品数推移", y_title="納品数（年合計）")
    amount_chart = _build_line_chart(title="金額推移", y_title="金額（円・年合計）")
    categories_qty = Reference(ws_src, min_col=1, min_row=2, max_row=row_count + 1)
    categories_amt = Reference(ws_src, min_col=7, min_row=2, max_row=row_count + 1)

    qty_series = {
        2: "実績",
        3: "直線延長予測",
        4: "重み付き回帰予測",
        5: "外部要因予測",
    }
    amount_series = {
        8: "実績",
        9: "直線延長予測",
        10: "重み付き回帰予測",
        11: "外部要因予測",
    }

    for col_idx, kind in qty_series.items():
        qty_chart.add_data(
            Reference(
                ws_src,
                min_col=col_idx,
                min_row=1,
                max_row=row_count + 1,
            ),
            titles_from_data=True,
        )
        qty_chart.set_categories(categories_qty)
        style = CHART_STYLES[kind]
        _configure_line_series(
            qty_chart.series[-1],
            color=style["color"],
            marker=style["marker"],
            dashed=style["dashed"],
        )

    for col_idx, kind in amount_series.items():
        amount_chart.add_data(
            Reference(
                ws_src,
                min_col=col_idx,
                min_row=1,
                max_row=row_count + 1,
            ),
            titles_from_data=True,
        )
        amount_chart.set_categories(categories_amt)
        style = CHART_STYLES[kind]
        _configure_line_series(
            amount_chart.series[-1],
            color=style["color"],
            marker=style["marker"],
            dashed=style["dashed"],
        )

    ws.add_chart(qty_chart, f"A{qty_chart_row}")
    ws.add_chart(amount_chart, f"A{amount_chart_row}")


def _write_forecast_chart_source(ws, comparison_df: pd.DataFrame) -> None:
    chart_source = pd.DataFrame(
        {
            "年": comparison_df.get("年"),
            "実績": comparison_df.get("実績\n納品数"),
            "直線": comparison_df.get("直線延長\n納品数"),
            "重み": comparison_df.get("重み付き回帰\n納品数"),
            "外部": comparison_df.get("外部要因予測\n納品数"),
            " ": [None] * len(comparison_df.index),
            "年 ": comparison_df.get("年"),
            "実績 ": comparison_df.get("実績\n金額"),
            "直線 ": comparison_df.get("直線延長\n金額"),
            "重み ": comparison_df.get("重み付き回帰\n金額"),
            "外部 ": comparison_df.get("外部要因予測\n金額"),
        }
    )
    for row_idx, row in enumerate(dataframe_to_rows(chart_source, index=False, header=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _auto_fit_columns(ws, *, max_col: int | None = None) -> None:
    limit = max_col or ws.max_column
    for col_idx in range(1, limit + 1):
        width = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            width = max(width, min(len(str(value)) * 1.3 + 2, 28))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _save_workbook(wb: Workbook, path: Path) -> None:
    try:
        wb.save(str(path))
    except OSError as e:
        raise ExportError(
            f"Excel ファイルの保存に失敗しました。アプリで開いていないか確認してください: {e}"
        ) from e
    except Exception as e:  # noqa: BLE001
        raise ExportError(f"Excel 出力中にエラーが発生しました: {e}") from e
